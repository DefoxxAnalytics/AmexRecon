"""
Amex → Zapro Supplier Reconciliation  |  Streamlit App
=======================================================
Solution 3 of 3 — Browser-based reconciliation tool.

Features:
  • Login / session management
  • Upload Amex XLS + Suppliers JSON + Invoices JSON + POs JSON
  • Configurable match threshold + live alias table editor
  • Colour-coded results table with summary KPI cards
  • Invoice/PO/Client enrichment from Zapro data exports
  • One-click download of formatted reconciliation Excel (3 sheets)

Run:
    streamlit run app.py
"""

import hashlib
import io
import json
import re
import tempfile
from datetime import datetime
from pathlib import Path

from fetch_zapro_data import ZaproClient, ZaproAPIError

import pandas as pd
import streamlit as st
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

# ── Page config (must be first Streamlit call) ────────────────────────────
st.set_page_config(
    page_title="Amex Reconciliation",
    page_icon="💳",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS & DEFAULTS
# ─────────────────────────────────────────────────────────────────────────────

VALID_USERS = {
    "admin": "foxx2026",
    "finance": "recon123",
}

DEFAULT_ALIASES = [
    {"From": "SQSP",        "To": "Squarespace"},
    {"From": "AMZN",        "To": "Amazon"},
    {"From": "HOMEDEPOT",   "To": "home depot"},
    {"From": "HOME DEPOT",  "To": "home depot"},
    {"From": "B2B PRIME",   "To": "Amazon"},
    {"From": "DISCOUNTTO",  "To": "DiscountToday"},
    {"From": "SP TRUDOOR",  "To": "Trudoor"},
    {"From": "LOWES",       "To": "Lowes"},
    {"From": "STAPLES",     "To": "Staples"},
    {"From": "RAPIDAPI",    "To": "RapidAPI"},
    {"From": "AUTH0",       "To": "Auth0"},
    {"From": "JOTFORM",     "To": "JotForm"},
    {"From": "NMSDC",       "To": "NMSDC"},
]

US_STATES = {
    "AL","AK","AZ","AR","CA","CO","CT","DE","FL","GA","HI","ID","IL","IN",
    "IA","KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV",
    "NH","NJ","NM","NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN",
    "TX","UT","VT","VA","WA","WV","WI","WY","DC",
}

COLOURS = {
    "AUTO MATCH": "#C6EFCE",
    "REVIEW":     "#FFEB9C",
    "NOT FOUND":  "#FFC7CE",
}

XLSX_COLOURS = {
    "AUTO MATCH": "C6EFCE",
    "REVIEW":     "FFEB9C",
    "NOT FOUND":  "FFC7CE",
    "header":     "1F3864",
    "subheader":  "2F75B6",
}

AMEX_DESC_COL  = 6
SKIP_KEYWORDS  = {"REMITTANCE", "PAYMENT", "BALANCE"}

# Supplier families — invoices for V1018 Amazon may be under V1024 Amazon
# Business Partner. Amount matching works cross-supplier within these groups.
SUPPLIER_GROUPS = [
    {"V1018", "V1024"},   # Amazon + Amazon Business Partner
    {"V1013", "V1002"},   # Menards duplicates
    {"V1008", "V1100"},   # H Hafner & Sons variants
]

NEAR_TOL_PCT = 0.02    # 2% tolerance for near-amount matching
NEAR_TOL_MIN = 0.50    # 50-cent floor

CONFIG_PATH = Path(__file__).parent / "config.json"
DEFAULT_CONFIG = {
    "zapro_base_url": "https://versatex.zapro.ai",
    "zapro_api_key": "",
}


def load_config():
    if CONFIG_PATH.exists():
        try:
            return {**DEFAULT_CONFIG, **json.loads(CONFIG_PATH.read_text(encoding="utf-8"))}
        except (json.JSONDecodeError, OSError):
            pass
    return dict(DEFAULT_CONFIG)


def save_config(base_url, api_key):
    CONFIG_PATH.write_text(
        json.dumps({"zapro_base_url": base_url, "zapro_api_key": api_key}, indent=2),
        encoding="utf-8",
    )


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────────────────────

def inject_css():
    st.markdown("""
    <style>
    /* ── Fonts ─────────────────────────────────────────────────────── */
    @import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display&family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

    /* ── Root palette ───────────────────────────────────────────────── */
    :root {
        --navy:   #1F3864;
        --blue:   #2F75B6;
        --lblue:  #E8F1FA;
        --green:  #C6EFCE;
        --amber:  #FFEB9C;
        --red:    #FFC7CE;
        --ink:    #1a1a2e;
        --muted:  #6B7280;
        --border: #E5E7EB;
        --bg:     #F8FAFC;
    }

    /* ── Global ─────────────────────────────────────────────────────── */
    html, body, [class*="css"] {
        font-family: 'DM Sans', sans-serif;
        background-color: var(--bg);
        color: var(--ink);
    }

    /* ── Hide default Streamlit chrome ──────────────────────────────── */
    #MainMenu, footer, header { visibility: hidden; }
    .block-container { padding-top: 1.5rem; padding-bottom: 2rem; }

    /* ── Login card ─────────────────────────────────────────────────── */
    .login-wrap {
        max-width: 420px;
        margin: 6rem auto 0;
        background: white;
        border-radius: 16px;
        padding: 3rem;
        box-shadow: 0 4px 32px rgba(31,56,100,.10);
        border: 1px solid var(--border);
    }
    .login-logo {
        font-family: 'DM Serif Display', serif;
        font-size: 1.9rem;
        color: var(--navy);
        margin-bottom: .25rem;
    }
    .login-sub {
        font-size: .85rem;
        color: var(--muted);
        margin-bottom: 2rem;
    }

    /* ── Top header bar ─────────────────────────────────────────────── */
    .app-header {
        display: flex;
        align-items: center;
        gap: 1rem;
        padding: .75rem 1.5rem;
        background: var(--navy);
        border-radius: 12px;
        margin-bottom: 1.5rem;
    }
    .app-header-title {
        font-family: 'DM Serif Display', serif;
        font-size: 1.35rem;
        color: white;
        flex: 1;
    }
    .app-header-user {
        font-size: .8rem;
        color: rgba(255,255,255,.65);
        font-family: 'DM Mono', monospace;
    }

    /* ── KPI cards ──────────────────────────────────────────────────── */
    .kpi-grid { display: flex; gap: 1rem; margin-bottom: 1.25rem; }
    .kpi-card {
        flex: 1;
        background: white;
        border-radius: 12px;
        padding: 1.1rem 1.25rem;
        border: 1px solid var(--border);
        box-shadow: 0 1px 4px rgba(0,0,0,.05);
    }
    .kpi-label {
        font-size: .72rem;
        font-weight: 600;
        letter-spacing: .08em;
        text-transform: uppercase;
        color: var(--muted);
        margin-bottom: .3rem;
    }
    .kpi-value {
        font-family: 'DM Serif Display', serif;
        font-size: 2.1rem;
        color: var(--navy);
        line-height: 1;
    }
    .kpi-sub {
        font-size: .75rem;
        color: var(--muted);
        margin-top: .2rem;
    }
    .kpi-green  .kpi-value { color: #22863a; }
    .kpi-amber  .kpi-value { color: #b45309; }
    .kpi-red    .kpi-value { color: #c0392b; }
    .kpi-navy   .kpi-value { color: var(--navy); }

    /* ── Section card ───────────────────────────────────────────────── */
    .section-card {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        border: 1px solid var(--border);
        margin-bottom: 1.25rem;
        box-shadow: 0 1px 4px rgba(0,0,0,.04);
    }
    .section-title {
        font-weight: 600;
        font-size: .9rem;
        color: var(--navy);
        letter-spacing: .04em;
        text-transform: uppercase;
        margin-bottom: 1rem;
        padding-bottom: .5rem;
        border-bottom: 2px solid var(--lblue);
    }

    /* ── Status badges ──────────────────────────────────────────────── */
    .badge {
        display: inline-block;
        padding: .2rem .65rem;
        border-radius: 99px;
        font-size: .72rem;
        font-weight: 600;
        letter-spacing: .04em;
    }
    .badge-green { background: #C6EFCE; color: #1a6b2a; }
    .badge-amber { background: #FFEB9C; color: #7c5c00; }
    .badge-red   { background: #FFC7CE; color: #8b0000; }

    /* ── Results table tweaks ───────────────────────────────────────── */
    .stDataFrame { border-radius: 10px; overflow: hidden; }
    .stDataFrame thead tr th {
        background: var(--navy) !important;
        color: white !important;
        font-family: 'DM Sans', sans-serif !important;
        font-size: .8rem !important;
        font-weight: 600 !important;
    }
    .stDataFrame tbody tr td {
        font-family: 'DM Mono', monospace !important;
        font-size: .8rem !important;
    }

    /* ── Sidebar ────────────────────────────────────────────────────── */
    section[data-testid="stSidebar"] {
        background: white;
        border-right: 1px solid var(--border);
    }
    section[data-testid="stSidebar"] .css-1d391kg { padding-top: 1.5rem; }

    /* ── Sidebar nav items ──────────────────────────────────────────── */
    .nav-item {
        display: flex;
        align-items: center;
        gap: .65rem;
        padding: .55rem .9rem;
        border-radius: 8px;
        margin-bottom: .2rem;
        cursor: pointer;
        font-size: .88rem;
        font-weight: 500;
        color: var(--ink);
        transition: background .15s;
    }
    .nav-item:hover, .nav-item.active {
        background: var(--lblue);
        color: var(--navy);
    }
    .nav-icon { font-size: 1.1rem; }

    /* ── Unmatched vendor chips ─────────────────────────────────────── */
    .chip-grid { display: flex; flex-wrap: wrap; gap: .4rem; margin-top: .5rem; }
    .chip {
        background: #FEE2E2;
        color: #7f1d1d;
        border-radius: 6px;
        padding: .2rem .6rem;
        font-size: .75rem;
        font-family: 'DM Mono', monospace;
        border: 1px solid #FCA5A5;
    }

    /* ── Upload zone ────────────────────────────────────────────────── */
    [data-testid="stFileUploader"] {
        border: 2px dashed var(--border) !important;
        border-radius: 10px !important;
        background: var(--bg) !important;
    }

    /* ── Button overrides ───────────────────────────────────────────── */
    .stButton > button[kind="primary"] {
        background: var(--navy);
        color: white;
        border: none;
        border-radius: 8px;
        font-family: 'DM Sans', sans-serif;
        font-weight: 600;
        letter-spacing: .04em;
        padding: .55rem 1.5rem;
        transition: background .15s, transform .1s;
    }
    .stButton > button[kind="primary"]:hover {
        background: var(--blue);
        transform: translateY(-1px);
    }

    /* ── Slider ─────────────────────────────────────────────────────── */
    .stSlider [data-baseweb="slider"] { margin-top: .25rem; }

    /* ── Progress bar ───────────────────────────────────────────────── */
    .stProgress > div > div { background: var(--blue); }

    /* ── Tab strip ──────────────────────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] { gap: .5rem; }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px 8px 0 0;
        font-family: 'DM Sans', sans-serif;
        font-weight: 500;
        font-size: .88rem;
    }
    .stTabs [aria-selected="true"] {
        background: white;
        color: var(--navy);
        font-weight: 600;
    }
    </style>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# MATCHING ENGINE  (inline — no import needed)
# ─────────────────────────────────────────────────────────────────────────────

def normalise(raw: str) -> str:
    s = raw.upper()
    s = re.sub(r'\*\S+', '', s)
    s = re.sub(r'#\S+', '', s)
    s = re.sub(r'\b(?:COM|NET|ORG|IO)\b', '', s)
    s = re.sub(r'\.(?:COM|NET|ORG|IO)', '', s)
    s = re.sub(r'\b\d{3}[-.\s]\d{3,4}[-.\s]\d{4}\b', '', s)
    s = re.sub(r'\b\d{4,}\b', '', s)
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    s = re.sub(r'\s{2,}', ' ', s).strip()
    tokens = s.split()
    while tokens and tokens[-1] in US_STATES:
        tokens.pop()
    uk_noise = {"LON", "GREATER", "LONDON"}
    while tokens and tokens[-1] in uk_noise:
        tokens.pop()
    return ' '.join(tokens[:5]).strip().lower()


def apply_alias(norm: str, alias_map: dict) -> str | None:
    upper = norm.upper()
    for prefix, canonical in alias_map.items():
        if upper.startswith(prefix.upper()):
            return canonical
    return None


def build_supplier_index(suppliers: list[dict]) -> dict[str, dict]:
    index = {}
    for s in sorted(suppliers, key=lambda x: x["display_identifier"]):
        if s.get("status") == "active":
            key = normalise(s["name"])
            if key not in index:
                index[key] = s
    return index


# ─────────────────────────────────────────────────────────────────────────────
# ENRICHMENT ENGINE  (Invoice + PO lookup)
# ─────────────────────────────────────────────────────────────────────────────

def build_invoice_indexes(invoices):
    from collections import defaultdict
    by_amount = defaultdict(list)
    by_vid    = defaultdict(list)
    for inv in invoices:
        try:
            amt = round(float(inv["invoice_net_total"]), 2)
            by_amount[amt].append(inv)
        except (ValueError, TypeError):
            pass
        vid = inv.get("supplier", {}).get("display_identifier")
        if vid:
            by_vid[vid].append(inv)
    return dict(by_amount), dict(by_vid)


def build_po_index(pos):
    return {p["display_identifier"]: p for p in pos}


def _get_project(inv, po_rec=None):
    for li in inv.get("line_items", []):
        for seg in li.get("billing_segments", []):
            if seg.get("segment_name") == "Project-Foxx" and seg.get("segment_value"):
                return seg["segment_value"]

    if po_rec:
        ship_title = (po_rec.get("ship_to_info") or {}).get("title", "")
        if ship_title:
            return ship_title

        for cf in po_rec.get("custom_fields") or []:
            name = (cf.get("field_name") or "").lower()
            if "project" in name or "client" in name:
                if cf.get("value"):
                    return cf["value"]

        bill_title = (po_rec.get("bill_to_info") or {}).get("title", "")
        if bill_title:
            return bill_title

    return ""


def _get_cf(record, field_name):
    for cf in record.get("custom_fields", []):
        if cf.get("field_name") == field_name:
            return cf.get("value") or ""
    return ""


def _fmt_amt(value):
    try:
        return f"${float(value):,.2f}"
    except (ValueError, TypeError):
        return ""


def _suppliers_related(vid_a, vid_b):
    for group in SUPPLIER_GROUPS:
        if vid_a in group and vid_b in group:
            return True
    return False


def enrich_transaction(supplier_vid, amex_amount, inv_by_amount, inv_by_vid, po_index):
    empty = dict(inv_match_type="NOT MATCHED", invoice_number="",
                 invoice_status="", invoice_net_total="", invoice_date="",
                 po_number="", po_net_total="", procore_po_id="", client_project="")

    if amex_amount is None:
        return {**empty, "inv_match_type": "NO AMOUNT"}

    amt = round(amex_amount, 2)

    exact_all = inv_by_amount.get(amt, [])
    same_sup  = [i for i in exact_all
                 if i["supplier"]["display_identifier"] == supplier_vid
                 or _suppliers_related(supplier_vid, i["supplier"]["display_identifier"])]
    if same_sup:
        n = len(same_sup)
        return _build_enrichment(same_sup[0], po_index,
                                 f"EXACT ({n} inv)" if n > 1 else "EXACT")

    if len(exact_all) == 1:
        return _build_enrichment(exact_all[0], po_index, "EXACT (amt only)")
    if len(exact_all) > 1:
        return _build_enrichment(exact_all[0], po_index,
                                 f"AMBIGUOUS ({len(exact_all)} inv)")

    candidates = list(inv_by_vid.get(supplier_vid, []))
    for group in SUPPLIER_GROUPS:
        if supplier_vid in group:
            for other in group:
                if other != supplier_vid:
                    candidates += inv_by_vid.get(other, [])
    tol  = max(amt * NEAR_TOL_PCT, NEAR_TOL_MIN)
    near = sorted(
        [c for c in candidates
         if abs(round(float(c["invoice_net_total"]), 2) - amt) <= tol],
        key=lambda x: abs(round(float(x["invoice_net_total"]), 2) - amt)
    )
    if near:
        return _build_enrichment(near[0], po_index, "NEAR MATCH")
    return empty


def _build_enrichment(inv, po_index, match_type):
    po_id  = (inv.get("po_details") or {}).get("display_identifier", "")
    po_rec = po_index.get(po_id) if po_id else None
    return dict(
        inv_match_type    = match_type,
        invoice_number    = inv.get("number", ""),
        invoice_status    = inv.get("status", ""),
        invoice_net_total = _fmt_amt(inv.get("invoice_net_total")),
        invoice_date      = (inv.get("invoice_date") or "")[:10],
        po_number         = po_id,
        po_net_total      = _fmt_amt(po_rec["po_net_total"]) if po_rec else "",
        procore_po_id     = _get_cf(inv, "Procore PO ID"),
        client_project    = _get_project(inv, po_rec),
    )


def run_matching(transactions, supplier_index, alias_map, auto_thresh, review_thresh):
    results = []
    for txn in transactions:
        norm    = normalise(txn["raw_merchant"])
        alias   = apply_alias(norm, alias_map)
        query   = normalise(alias) if alias else norm

        top_matches = process.extract(query, supplier_index.keys(),
                                       scorer=fuzz.token_set_ratio,
                                       score_cutoff=review_thresh, limit=3)
        if top_matches:
            matched_key, score, _ = top_matches[0]
            supplier = supplier_index[matched_key]
        else:
            supplier, score = None, 0

        score = int(score)
        if score >= auto_thresh:
            status = "AUTO MATCH"
        elif score >= review_thresh:
            status = "REVIEW"
        else:
            status = "NOT FOUND"

        alt_matches = ""
        if status == "REVIEW" and len(top_matches) > 1:
            alts = [f"{supplier_index[k]['name']} ({int(s)})"
                    for k, s, _ in top_matches[1:]]
            alt_matches = " | ".join(alts)

        results.append({
            **txn,
            "normalised":   norm,
            "alias_used":   alias or "",
            "matched_name": supplier["name"]               if supplier else "",
            "supplier_id":  supplier["display_identifier"] if supplier else "",
            "score":        score,
            "status":       status,
            "alt_matches":  alt_matches,
        })
    return results


# ─────────────────────────────────────────────────────────────────────────────
# FILE LOADERS
# ─────────────────────────────────────────────────────────────────────────────

def load_amex_bytes(file_bytes: bytes) -> list[dict]:
    wb = xlrd.open_workbook(file_contents=file_bytes)
    sh = wb.sheets()[0]
    txns = []
    for r in range(1, sh.nrows):
        row  = sh.row_values(r)
        desc = str(row[AMEX_DESC_COL]).strip()
        if not desc:
            continue
        if any(kw in desc.upper() for kw in SKIP_KEYWORDS):
            continue
        txns.append({
            "row_num":      r + 1,
            "cardmember":   str(row[1]).strip(),
            "proc_date":    str(row[2]).strip(),
            "txn_date":     str(row[3]).strip(),
            "ref_no":       str(row[4]).strip(),
            "amount_usd":   str(row[5]).strip(),
            "raw_merchant": desc,
        })
    return txns


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(results: list[dict], statement_name: str) -> bytes:
    wb   = Workbook()
    thin = Side(style="thin", color="BFBFBF")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style(ws, r, c, val, fill=None, bold=False, align="left",
              font_color=None, italic=False):
        cell = ws.cell(row=r, column=c, value=val)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        white = fill in (XLSX_COLOURS["header"], XLSX_COLOURS["subheader"])
        fc = font_color or ("FFFFFF" if (bold and white) else "000000")
        cell.font = Font(bold=bold, italic=italic, color=fc, size=10, name="Arial")
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
        cell.border = bdr
        return cell

    # ── Column definitions ────────────────────────────────────────────
    # (header, data_key, width, group)  group: "core" | "enrich"
    COLS = [
        ("Row#",              "row_num",          6,   "core"),
        ("Cardmember",        "cardmember",        14,  "core"),
        ("Proc Date",         "proc_date",         12,  "core"),
        ("Txn Date",          "txn_date",          12,  "core"),
        ("Ref No",            "ref_no",            20,  "core"),
        ("Amount (USD)",      "amount_usd",        13,  "core"),
        ("Raw Amex Merchant", "raw_merchant",      44,  "core"),
        ("Normalised",        "normalised",        28,  "core"),
        ("Alias Used",        "alias_used",        16,  "core"),
        ("Matched Supplier",  "matched_name",      32,  "core"),
        ("Supplier ID",       "supplier_id",       11,  "core"),
        ("Match Score",       "score",             11,  "core"),
        ("Match Status",      "status",            13,  "core"),
        ("Invoice #",         "invoice_number",    28,  "enrich"),
        ("Invoice Date",      "invoice_date",      13,  "enrich"),
        ("Invoice Status",    "invoice_status",    14,  "enrich"),
        ("Invoice Net Total", "invoice_net_total", 16,  "enrich"),
        ("PO Number",         "po_number",         10,  "enrich"),
        ("PO Net Total",      "po_net_total",      13,  "enrich"),
        ("Procore PO ID",     "procore_po_id",     22,  "enrich"),
        ("Client / Project",  "client_project",    42,  "enrich"),
        ("Inv Match Type",    "inv_match_type",    18,  "enrich"),
    ]

    enriched_hdr  = "E2EFDA"   # soft green for enrichment header cells
    no_inv_fill   = "F2F2F2"   # grey for rows with no invoice match
    inv_match_fill = "D5E8D4"  # green for matched invoice amount cell

    # ── Sheet 1: Reconciliation ───────────────────────────────────────
    ws = wb.active
    ws.title = "Reconciliation"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    for c_idx, (label, _, width, group) in enumerate(COLS, 1):
        fill = enriched_hdr           if group == "enrich" else XLSX_COLOURS["header"]
        fc   = XLSX_COLOURS["header"] if group == "enrich" else "FFFFFF"
        style(ws, 1, c_idx, label, fill=fill, bold=True, align="center", font_color=fc)
        ws.column_dimensions[get_column_letter(c_idx)].width = width

    has_enrichment = any(r.get("inv_match_type") not in ("", None)
                         for r in results)

    for ri, rec in enumerate(results, 2):
        status_val = rec.get("status", "")
        row_fill   = XLSX_COLOURS.get(status_val, "FFFFFF")
        inv_type   = rec.get("inv_match_type", "")
        no_inv     = inv_type in ("", "NOT MATCHED", "NO AMOUNT")

        for c_idx, (_, key, _, group) in enumerate(COLS, 1):
            val   = rec.get(key, "") or ""
            align = ("center" if key in ("row_num","score","status","inv_match_type",
                                          "invoice_status","proc_date","txn_date",
                                          "invoice_date","po_number","supplier_id")
                     else "right" if key in ("amount_usd","invoice_net_total","po_net_total")
                     else "left")
            # Enrichment cols grey out when no invoice found
            cell_fill = no_inv_fill if (group == "enrich" and no_inv and has_enrichment) else None
            bold      = key == "status"
            style(ws, ri, c_idx, val, fill=cell_fill, bold=bold, align=align)

        # Status cell coloured by match result
        style(ws, ri, 13, status_val, fill=row_fill, bold=True, align="center")
        # Invoice net total green when matched
        inv_col = next(i for i,(_, k, _, _) in enumerate(COLS, 1) if k == "invoice_net_total")
        if rec.get("invoice_net_total"):
            style(ws, ri, inv_col, rec["invoice_net_total"],
                  fill=inv_match_fill, align="right")

    # ── Sheet 2: Summary ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    auto_r   = [r for r in results if r.get("status") == "AUTO MATCH"]
    review_r = [r for r in results if r.get("status") == "REVIEW"]
    nf_r     = [r for r in results if r.get("status") == "NOT FOUND"]
    inv_hit  = [r for r in results
                if r.get("inv_match_type") not in ("", "NOT MATCHED", "NO AMOUNT", None)]

    total_amt = 0.0
    for r in results:
        try:
            total_amt += float(str(r.get("amount_usd","")).replace(",",""))
        except (ValueError, TypeError):
            pass

    def s_row(ws, ri, label, val, detail, is_hdr=False, fill=None):
        for ci, v in enumerate([label, val, detail], 1):
            style(ws2, ri, ci, "" if is_hdr and ci > 1 else v,
                  fill=fill, bold=is_hdr or ci == 1,
                  font_color="FFFFFF" if is_hdr else ("555555" if ci == 3 else None),
                  italic=(ci == 3 and not is_hdr))

    rows = [
        ("SUPPLIER MATCHING",   None,        None,                True,  XLSX_COLOURS["header"]),
        ("Total transactions",  len(results),"",                  False, None),
        ("Total spend (USD)",   f"${total_amt:,.2f}", "",         False, None),
        ("Auto-matched",        len(auto_r), "Score ≥ 75",        False, None),
        ("Needs review",        len(review_r),"Score 50–74",      False, None),
        ("Not found",           len(nf_r),   "Score < 50",        False, None),
        ("",                    "",          "",                   False, None),
        ("INVOICE ENRICHMENT",  None,        None,                True,  XLSX_COLOURS["subheader"]),
        ("Invoices matched",    len(inv_hit),f"of {len(results)}",False, None),
        ("Not matched",         len(results)-len(inv_hit),"",     False, None),
        ("",                    "",          "",                   False, None),
        ("UNMATCHED VENDORS",   None,        None,                True,  XLSX_COLOURS["subheader"]),
    ]
    for ri, (label, val, detail, is_hdr, fill) in enumerate(rows, 1):
        s_row(ws2, ri, label, val, detail, is_hdr=is_hdr, fill=fill)

    for i, m in enumerate(sorted({r["raw_merchant"] for r in nf_r}), len(rows)+1):
        ws2.cell(row=i, column=1, value=m).border = bdr
        style(ws2, i, 2, "Action required", italic=True, font_color="888888")

    for ci, w in zip([1,2,3], [46, 22, 38]):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Invoice Detail (matched rows only) ───────────────────
    if inv_hit:
        ws3 = wb.create_sheet("Invoice Detail")
        ws3.freeze_panes = "A2"
        ws3.row_dimensions[1].height = 28

        D_COLS = [
            ("Txn Date",          "txn_date",          12),
            ("Amount (USD)",      "amount_usd",        13),
            ("Raw Amex Merchant", "raw_merchant",       40),
            ("Matched Supplier",  "matched_name",       28),
            ("Supplier ID",       "supplier_id",        11),
            ("Invoice #",         "invoice_number",     28),
            ("Invoice Date",      "invoice_date",       13),
            ("Invoice Status",    "invoice_status",     14),
            ("Invoice Net Total", "invoice_net_total",  16),
            ("PO Number",         "po_number",          10),
            ("PO Net Total",      "po_net_total",       13),
            ("Procore PO ID",     "procore_po_id",      22),
            ("Client / Project",  "client_project",     44),
            ("Inv Match Type",    "inv_match_type",     18),
        ]
        for c_idx, (label, _, width) in enumerate(D_COLS, 1):
            style(ws3, 1, c_idx, label, fill=XLSX_COLOURS["header"],
                  bold=True, align="center", font_color="FFFFFF")
            ws3.column_dimensions[get_column_letter(c_idx)].width = width

        for ri, rec in enumerate(inv_hit, 2):
            inv_type = rec.get("inv_match_type", "")
            row_bg   = ("F0FFF0" if inv_type.startswith("EXACT")
                        else "FFF9E6" if inv_type == "NEAR MATCH" else None)
            for c_idx, (_, key, _) in enumerate(D_COLS, 1):
                val   = rec.get(key, "") or ""
                align = ("right"  if key in ("amount_usd","invoice_net_total","po_net_total")
                         else "center" if key in ("txn_date","invoice_date","supplier_id",
                                                   "po_number","invoice_status","inv_match_type")
                         else "left")
                style(ws3, ri, c_idx, val, fill=row_bg, align=align)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────────────────────────────────────

def init_state():
    defaults = {
        "logged_in":       False,
        "username":        "",
        "results":         None,
        "suppliers":       None,
        "transactions":    None,
        "invoices":        None,
        "purchase_orders": None,
        "aliases":         DEFAULT_ALIASES.copy(),
        "auto_thresh":     75,
        "review_thresh":   50,
        "statement_name":  "Amex Statement",
        "active_tab":      "upload",
        "last_run_config": None,
        "zapro_base_url":  load_config()["zapro_base_url"],
        "zapro_api_key":   load_config()["zapro_api_key"],
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


# ─────────────────────────────────────────────────────────────────────────────
# PAGES
# ─────────────────────────────────────────────────────────────────────────────

def page_login():
    st.markdown("""
    <div class="login-wrap">
        <div class="login-logo">💳 AmexRecon</div>
        <div class="login-sub">Zapro Supplier Reconciliation Portal</div>
    </div>
    """, unsafe_allow_html=True)

    col_a, col_b, col_c = st.columns([1, 1.4, 1])
    with col_b:
        st.markdown("<div style='margin-top:-15.5rem'>", unsafe_allow_html=True)
        with st.form("login_form"):
            st.markdown("<div style='height:11rem'></div>", unsafe_allow_html=True)
            username = st.text_input("Username", placeholder="admin")
            password = st.text_input("Password", type="password", placeholder="••••••••")
            submitted = st.form_submit_button("Sign In", use_container_width=True, type="primary")
            if submitted:
                if VALID_USERS.get(username) == password:
                    st.session_state.logged_in = True
                    st.session_state.username  = username
                    st.rerun()
                else:
                    st.error("Invalid credentials. Try admin / foxx2026")
        st.markdown("</div>", unsafe_allow_html=True)
        st.markdown(
            "<p style='text-align:center;font-size:.75rem;color:#9CA3AF;margin-top:.5rem'>"
            "Demo credentials: admin / foxx2026</p>",
            unsafe_allow_html=True
        )


def render_header():
    st.markdown(f"""
    <div class="app-header">
        <div class="app-header-title">💳 Amex → Zapro Reconciliation</div>
        <div class="app-header-user">Signed in as <strong>{st.session_state.username}</strong></div>
    </div>
    """, unsafe_allow_html=True)


def render_sidebar():
    with st.sidebar:
        st.markdown("### ⚙️ Settings")
        st.markdown("---")

        st.markdown("**Match Thresholds**")
        auto_t = st.slider("Auto-match floor", 60, 95,
                           st.session_state.auto_thresh, 5,
                           help="Score ≥ this → AUTO MATCH")
        review_t = st.slider("Review floor", 30, int(auto_t) - 5,
                              min(st.session_state.review_thresh, int(auto_t) - 5), 5,
                              help="Score between this and auto floor → REVIEW")
        st.session_state.auto_thresh   = auto_t
        st.session_state.review_thresh = review_t

        st.markdown("---")
        st.markdown("**Alias Table**")
        st.caption("Maps Amex billing codes to supplier names before matching.")

        alias_df = pd.DataFrame(st.session_state.aliases)
        edited = st.data_editor(
            alias_df,
            use_container_width=True,
            num_rows="dynamic",
            hide_index=True,
            column_config={
                "From": st.column_config.TextColumn("Amex Code", width="small"),
                "To":   st.column_config.TextColumn("Supplier Name", width="medium"),
            },
            key="alias_editor",
        )
        st.session_state.aliases = edited.dropna(how="all").to_dict("records")

        st.markdown("---")
        if st.session_state.username == "admin":
            st.markdown("**Zapro API Configuration**")
            new_url = st.text_input(
                "Base URL",
                value=st.session_state.zapro_base_url,
                key="zapro_url_input",
            )
            new_key = st.text_input(
                "API Key",
                value=st.session_state.zapro_api_key,
                type="password",
                key="zapro_key_input",
            )
            if st.button("Save API Config", use_container_width=True):
                st.session_state.zapro_base_url = new_url
                st.session_state.zapro_api_key = new_key
                save_config(new_url, new_key)
                st.success("API config saved")
        else:
            if st.session_state.zapro_api_key:
                st.success("Zapro API configured")
            else:
                st.caption("Zapro API not configured. Ask admin to set credentials.")

        st.markdown("---")
        if st.button("🚪 Sign Out", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()


def _fetch_zapro_data():
    client = ZaproClient(st.session_state.zapro_base_url, st.session_state.zapro_api_key)
    with st.status("Fetching data from Zapro API...", expanded=True) as status:
        st.write("Generating auth token...")
        client.generate_token()

        st.write("Fetching suppliers...")
        suppliers = client.fetch_suppliers()
        st.session_state.suppliers = suppliers
        active = sum(1 for s in suppliers if s.get("status") == "active")
        st.write(f"  {len(suppliers)} suppliers ({active} active)")

        st.write("Fetching invoices...")
        invoices = client.fetch_invoices()
        st.session_state.invoices = invoices
        st.write(f"  {len(invoices)} invoices")

        st.write("Fetching purchase orders...")
        pos = client.fetch_purchase_orders()
        st.session_state.purchase_orders = pos
        st.write(f"  {len(pos)} purchase orders")

        status.update(label="Zapro data loaded", state="complete", expanded=False)


def page_upload():
    st.markdown("### 📂 Upload Files")

    # ── Fetch from Zapro API ──────────────────────────────────────────
    _render_amex_uploader()

    if st.session_state.zapro_api_key:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Zapro Data</div>', unsafe_allow_html=True)
        api_col1, api_col2 = st.columns([3, 1])
        with api_col1:
            already_loaded = all([
                st.session_state.suppliers,
                st.session_state.invoices,
                st.session_state.purchase_orders,
            ])
            if already_loaded:
                sup_count = len(st.session_state.suppliers)
                inv_count = len(st.session_state.invoices)
                po_count  = len(st.session_state.purchase_orders)
                st.success(f"Zapro data loaded — {sup_count} suppliers, {inv_count} invoices, {po_count} POs")
            else:
                st.caption(f"Endpoint: {st.session_state.zapro_base_url}")
        with api_col2:
            if st.button("Fetch from Zapro", use_container_width=True, type="primary", key="fetch_zapro"):
                try:
                    _fetch_zapro_data()
                    st.rerun()
                except ZaproAPIError as exc:
                    st.error(f"Zapro API error: {exc}")
                except Exception as exc:
                    st.error(f"Failed to fetch: {exc}")
        st.markdown('</div>', unsafe_allow_html=True)
        if st.session_state.purchase_orders and st.session_state.username == "admin":
            with st.expander("🔍 DEBUG: PO Custom Fields (remove later)"):
                sample_pos = st.session_state.purchase_orders[:3]
                for po in sample_pos:
                    st.markdown(f"**PO: {po.get('display_identifier')}**")
                    top_cfs = po.get("custom_fields") or []
                    if top_cfs:
                        st.write("Top-level custom_fields:")
                        for cf in top_cfs:
                            st.write(f"  `{cf.get('field_name')}`: `{cf.get('value')}`")
                    else:
                        st.write("No top-level custom_fields")
                    for li in (po.get("line_items") or [])[:2]:
                        li_cfs = li.get("custom_fields") or []
                        if li_cfs:
                            st.write(f"Line item {li.get('line_number')} custom_fields:")
                            for cf in li_cfs:
                                st.write(f"  `{cf.get('field_name')}`: `{cf.get('value')}`")
                    st.markdown("---")
        with st.expander("Or upload JSON files manually"):
            _render_zapro_uploaders()
    else:
        _render_zapro_uploaders()

    # ── Run button ────────────────────────────────────────────────────
    st.markdown("---")
    ready = st.session_state.transactions and st.session_state.suppliers
    enrich_ready = st.session_state.invoices and st.session_state.purchase_orders
    if not ready:
        st.info("Upload or load the Amex XLS and Suppliers files to enable matching.")
    elif not enrich_ready:
        st.info("Load Invoices and Purchase Orders files to enable invoice/PO/client enrichment.")

    if st.button("🔍 Run Matching", type="primary",
                 disabled=not ready, use_container_width=False):
        run_and_store()


def _render_amex_uploader():
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Amex Statement (.xls)</div>', unsafe_allow_html=True)
    amex_file = st.file_uploader(
        "Drop Amex XLS here", type=["xls"],
        label_visibility="collapsed", key="amex_upload"
    )
    if amex_file:
        try:
            st.session_state.statement_name = amex_file.name.replace(".xls","")
            txns = load_amex_bytes(amex_file.read())
            st.session_state.transactions = txns
            st.success(f"✅  {len(txns)} transactions loaded")
        except Exception as exc:
            st.error(f"Failed to parse Amex file: {exc}")
    elif st.button("Use sample file", key="use_sample_amex"):
        sample = Path("/mnt/user-data/uploads/Amex_test.xls")
        if sample.exists():
            txns = load_amex_bytes(sample.read_bytes())
            st.session_state.transactions = txns
            st.session_state.statement_name = "Statement_1008_Feb_2026"
            st.success(f"✅  {len(txns)} transactions loaded from sample")
        else:
            st.warning("Sample file not found.")
    st.markdown('</div>', unsafe_allow_html=True)


def _render_zapro_uploaders():
    c1, c2 = st.columns(2)
    with c1:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Zapro Suppliers (.json)</div>', unsafe_allow_html=True)
        sup_file = st.file_uploader(
            "Drop suppliers JSON here", type=["json"],
            label_visibility="collapsed", key="sup_upload"
        )
        if sup_file:
            try:
                sup_data = json.load(sup_file)
                st.session_state.suppliers = sup_data
                active = sum(1 for s in sup_data if s.get("status") == "active")
                st.success(f"✅  {len(sup_data)} suppliers loaded ({active} active)")
            except (json.JSONDecodeError, Exception) as exc:
                st.error(f"Failed to parse suppliers JSON: {exc}")
        elif st.button("Use sample file", key="use_sample_sup"):
            sample = Path("/mnt/user-data/uploads/suppliers.json")
            if sample.exists():
                sup_data = json.loads(sample.read_text())
                st.session_state.suppliers = sup_data
                active = sum(1 for s in sup_data if s.get("status") == "active")
                st.success(f"✅  {len(sup_data)} suppliers loaded ({active} active) from sample")
            else:
                st.warning("Sample file not found.")
        st.markdown('</div>', unsafe_allow_html=True)

    with c2:
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">Zapro Invoices (.json) <span style="font-weight:400;color:#6B7280">— enrichment</span></div>', unsafe_allow_html=True)
        inv_file = st.file_uploader(
            "Drop invoices JSON here", type=["json"],
            label_visibility="collapsed", key="inv_upload"
        )
        if inv_file:
            try:
                inv_data = json.load(inv_file)
                st.session_state.invoices = inv_data
                st.success(f"✅  {len(inv_data)} invoices loaded")
            except (json.JSONDecodeError, Exception) as exc:
                st.error(f"Failed to parse invoices JSON: {exc}")
        elif st.button("Use sample file", key="use_sample_inv"):
            sample = Path("/mnt/user-data/uploads/invoices.json")
            if sample.exists():
                inv_data = json.loads(sample.read_text())
                st.session_state.invoices = inv_data
                st.success(f"✅  {len(inv_data)} invoices loaded from sample")
            else:
                st.warning("Sample file not found.")
        if not st.session_state.invoices:
            st.caption("Optional — skip to run supplier matching only.")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.markdown('<div class="section-title">Zapro Purchase Orders (.json) <span style="font-weight:400;color:#6B7280">— enrichment</span></div>', unsafe_allow_html=True)
    po_file = st.file_uploader(
        "Drop purchase_orders JSON here", type=["json"],
        label_visibility="collapsed", key="po_upload"
    )
    if po_file:
        try:
            po_data = json.load(po_file)
            st.session_state.purchase_orders = po_data
            st.success(f"✅  {len(po_data)} purchase orders loaded")
        except (json.JSONDecodeError, Exception) as exc:
            st.error(f"Failed to parse purchase orders JSON: {exc}")
    elif st.button("Use sample file", key="use_sample_po"):
        sample = Path("/mnt/user-data/uploads/purchase_orders.json")
        if sample.exists():
            po_data = json.loads(sample.read_text())
            st.session_state.purchase_orders = po_data
            st.success(f"✅  {len(po_data)} purchase orders loaded from sample")
        else:
            st.warning("Sample file not found.")
    if not st.session_state.purchase_orders:
        st.caption("Optional — required to show PO amounts and Procore IDs.")
    st.markdown('</div>', unsafe_allow_html=True)


def _config_hash():
    raw = json.dumps(st.session_state.aliases, sort_keys=True)
    raw += f"|{st.session_state.auto_thresh}|{st.session_state.review_thresh}"
    return hashlib.md5(raw.encode()).hexdigest()


def run_and_store():
    txns      = st.session_state.transactions
    suppliers = st.session_state.suppliers
    invoices  = st.session_state.invoices        or []
    pos       = st.session_state.purchase_orders or []

    alias_map = {row["From"]: row["To"] for row in st.session_state.aliases
                 if row.get("From") and row.get("To")}

    progress = st.progress(0, text="Building indexes…")

    # Supplier index
    sup_index = build_supplier_index(suppliers)

    # Invoice + PO indexes (empty dicts if files not loaded)
    if invoices:
        inv_by_amount, inv_by_vid = build_invoice_indexes(invoices)
    else:
        inv_by_amount, inv_by_vid = {}, {}

    po_index = build_po_index(pos) if pos else {}

    enriching = bool(invoices and pos)

    # DEBUG: surface enrichment state
    st.sidebar.markdown("---")
    st.sidebar.markdown("**DEBUG: Enrichment**")
    st.sidebar.write(f"invoices: {len(invoices)}, pos: {len(pos)}")
    st.sidebar.write(f"enriching: {enriching}")
    st.sidebar.write(f"inv_by_amount keys: {len(inv_by_amount)}")
    st.sidebar.write(f"inv_by_vid keys: {len(inv_by_vid)}")
    st.sidebar.write(f"po_index keys: {len(po_index)}")
    if invoices:
        sample = invoices[0]
        st.sidebar.write(f"Sample invoice keys: {list(sample.keys())[:10]}")
        st.sidebar.write(f"  invoice_net_total: {sample.get('invoice_net_total')}")
        st.sidebar.write(f"  supplier: {sample.get('supplier')}")
        st.sidebar.write(f"  po_details: {sample.get('po_details')}")
    if pos:
        sample_po = pos[0]
        st.sidebar.write(f"Sample PO keys: {list(sample_po.keys())[:10]}")
        st.sidebar.write(f"  display_identifier: {sample_po.get('display_identifier')}")
    # END DEBUG

    progress.progress(10, text=f"Matching {len(txns)} transactions…")
    results = []
    auto_t   = st.session_state.auto_thresh
    review_t = st.session_state.review_thresh

    for i, txn in enumerate(txns):
        # ── Supplier fuzzy match ───────────────────────────────────────
        norm    = normalise(txn["raw_merchant"])
        alias   = apply_alias(norm, alias_map)
        query   = normalise(alias) if alias else norm
        top_matches = process.extract(query, sup_index.keys(),
                                       scorer=fuzz.token_set_ratio,
                                       score_cutoff=review_t, limit=3)
        if top_matches:
            matched_key, score, _ = top_matches[0]
            supplier = sup_index[matched_key]
        else:
            supplier, score = None, 0

        score    = int(score)
        status   = ("AUTO MATCH" if score >= auto_t
                    else "REVIEW"    if score >= review_t
                    else "NOT FOUND")

        alt_matches = ""
        if status == "REVIEW" and len(top_matches) > 1:
            alts = [f"{sup_index[k]['name']} ({int(s)})"
                    for k, s, _ in top_matches[1:]]
            alt_matches = " | ".join(alts)

        vid = supplier["display_identifier"] if supplier else ""

        # ── Invoice + PO enrichment ────────────────────────────────────
        try:
            amex_amt = round(float(str(txn["amount_usd"]).replace(",", "")), 2)
        except (ValueError, TypeError):
            amex_amt = None

        enrichment = (enrich_transaction(vid, amex_amt, inv_by_amount, inv_by_vid, po_index)
                      if enriching else
                      dict(inv_match_type="", invoice_number="", invoice_status="",
                           invoice_net_total="", invoice_date="", po_number="",
                           po_net_total="", procore_po_id="", client_project=""))

        results.append({
            **txn,
            "normalised":   norm,
            "alias_used":   alias or "",
            "matched_name": supplier["name"]               if supplier else "",
            "supplier_id":  vid,
            "score":        score,
            "status":       status,
            "alt_matches":  alt_matches,
            **enrichment,
        })
        progress.progress(10 + int(88 * (i + 1) / len(txns)),
                          text=f"Matched {i+1}/{len(txns)}…")

    st.session_state.results = results
    progress.progress(100, text="Done!")
    inv_hits = sum(1 for r in results
                   if r.get("inv_match_type") not in ("", "NOT MATCHED", "NO AMOUNT"))
    enrich_note = f" | 📄 {inv_hits}/{len(results)} invoices matched" if enriching else ""
    st.success(f"✅  Complete — {len(results)} rows processed{enrich_note}")
    st.session_state.last_run_config = _config_hash()
    st.rerun()


def page_results():
    results    = st.session_state.results
    enriching  = any(r.get("inv_match_type") not in ("", None) for r in results)

    if (st.session_state.get("last_run_config")
            and _config_hash() != st.session_state.last_run_config):
        st.warning("⚠️ Settings changed since last run — click **Run Matching** to update results.")

    # ── KPI cards ─────────────────────────────────────────────────────
    auto_r   = [r for r in results if r["status"] == "AUTO MATCH"]
    review_r = [r for r in results if r["status"] == "REVIEW"]
    nf_r     = [r for r in results if r["status"] == "NOT FOUND"]
    inv_hit  = [r for r in results
                if r.get("inv_match_type") not in ("", "NOT MATCHED", "NO AMOUNT", None)]
    total_amt = 0.0
    for r in results:
        try:
            total_amt += float(str(r["amount_usd"]).replace(",",""))
        except (ValueError, TypeError):
            pass

    # Row 1 — supplier matching KPIs
    st.markdown(f"""
    <div class="kpi-grid">
        <div class="kpi-card kpi-navy">
            <div class="kpi-label">Total Transactions</div>
            <div class="kpi-value">{len(results)}</div>
            <div class="kpi-sub">{st.session_state.statement_name}</div>
        </div>
        <div class="kpi-card kpi-navy">
            <div class="kpi-label">Total Spend</div>
            <div class="kpi-value">${total_amt:,.0f}</div>
            <div class="kpi-sub">USD</div>
        </div>
        <div class="kpi-card kpi-green">
            <div class="kpi-label">Auto Matched</div>
            <div class="kpi-value">{len(auto_r)}</div>
            <div class="kpi-sub">Score ≥ {st.session_state.auto_thresh} — ready to post</div>
        </div>
        <div class="kpi-card kpi-amber">
            <div class="kpi-label">Needs Review</div>
            <div class="kpi-value">{len(review_r)}</div>
            <div class="kpi-sub">Score {st.session_state.review_thresh}–{st.session_state.auto_thresh - 1}</div>
        </div>
        <div class="kpi-card kpi-red">
            <div class="kpi-label">Not Found</div>
            <div class="kpi-value">{len(nf_r)}</div>
            <div class="kpi-sub">Add to suppliers.json</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # Row 2 — enrichment KPIs (only shown when enrichment data was loaded)
    if enriching:
        inv_pct  = int(len(inv_hit) / len(results) * 100) if results else 0
        no_inv_r = [r for r in results
                    if r.get("inv_match_type") in ("NOT MATCHED", "NO AMOUNT", "", None)]
        # Unique client/projects found
        projects = sorted({r.get("client_project","") for r in inv_hit
                           if r.get("client_project","")})
        proj_display = ", ".join(projects[:3]) + ("…" if len(projects) > 3 else "")

        st.markdown(f"""
        <div class="kpi-grid">
            <div class="kpi-card kpi-green">
                <div class="kpi-label">Invoices Matched</div>
                <div class="kpi-value">{len(inv_hit)}</div>
                <div class="kpi-sub">{inv_pct}% of transactions enriched</div>
            </div>
            <div class="kpi-card kpi-red">
                <div class="kpi-label">No Invoice Found</div>
                <div class="kpi-value">{len(no_inv_r)}</div>
                <div class="kpi-sub">SaaS / memberships / new vendors</div>
            </div>
            <div class="kpi-card kpi-navy">
                <div class="kpi-label">Client Projects</div>
                <div class="kpi-value">{len(projects)}</div>
                <div class="kpi-sub" title="{', '.join(projects)}">{proj_display or "—"}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── Unmatched vendor chips ─────────────────────────────────────────
    if nf_r:
        unmatched_names = sorted({r["raw_merchant"] for r in nf_r})
        chips = "".join(f'<span class="chip">{n[:40]}</span>' for n in unmatched_names)
        st.markdown(f"""
        <div class="section-card">
            <div class="section-title">❌ Vendors Not in Supplier List</div>
            <div class="chip-grid">{chips}</div>
        </div>
        """, unsafe_allow_html=True)

    # ── Results tabs ──────────────────────────────────────────────────
    tabs = ["All Results", "Needs Review / Not Found", "Auto Matched"]
    if enriching:
        tabs.append("📄 Invoice Detail")
    tab_objs = st.tabs(tabs)

    # Core columns always shown
    BASE_COLS  = ["row_num","txn_date","amount_usd","raw_merchant",
                  "matched_name","supplier_id","score","status","alt_matches"]
    BASE_NAMES = ["Row","Date","Amount","Amex Merchant",
                  "Matched Supplier","ID","Score","Status","Alternatives"]
    # Enrichment columns appended when data is available
    ENR_COLS   = ["invoice_number","invoice_date","invoice_status",
                  "invoice_net_total","po_number","po_net_total",
                  "procore_po_id","client_project","inv_match_type"]
    ENR_NAMES  = ["Invoice #","Inv Date","Inv Status",
                  "Invoice Total","PO #","PO Total",
                  "Procore PO ID","Client / Project","Inv Match"]

    show_cols  = BASE_COLS  + (ENR_COLS  if enriching else [])
    show_names = BASE_NAMES + (ENR_NAMES if enriching else [])

    def render_table(data):
        if not data:
            st.info("No rows in this category.")
            return
        df = pd.DataFrame(data)[show_cols].rename(
            columns=dict(zip(show_cols, show_names)))

        def colour_row(row):
            c = {"AUTO MATCH": "#e8f5e9", "REVIEW": "#fff8e1", "NOT FOUND": "#ffebee"}
            bg = c.get(row["Status"], "")
            return [f"background-color: {bg}"] * len(row)

        styled = df.style.apply(colour_row, axis=1)
        col_cfg = {"Score": st.column_config.ProgressColumn(
                       "Score", min_value=0, max_value=100, format="%d")}
        if enriching:
            col_cfg["Invoice Total"] = st.column_config.TextColumn("Invoice Total")
            col_cfg["PO Total"]      = st.column_config.TextColumn("PO Total")
        st.dataframe(styled, use_container_width=True, hide_index=True,
                     column_config=col_cfg)

    with tab_objs[0]: render_table(results)
    with tab_objs[1]: render_table([r for r in results if r["status"] in ("REVIEW","NOT FOUND")])
    with tab_objs[2]: render_table([r for r in results if r["status"] == "AUTO MATCH"])

    if enriching and len(tab_objs) > 3:
        with tab_objs[3]:
            if inv_hit:
                render_table(inv_hit)
            else:
                st.info("No invoices were matched.")

    # ── Download ──────────────────────────────────────────────────────
    st.markdown("---")
    xlsx_bytes = build_excel(results, st.session_state.statement_name)
    sheet_note = "Reconciliation + Summary + Invoice Detail" if enriching else "Reconciliation + Summary"
    fname = (f"amex_recon_{st.session_state.statement_name}_"
             f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    st.download_button(
        label="⬇️  Download Reconciliation Excel",
        data=xlsx_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
    st.caption(f"File: {fname}  •  {len(results)} rows  •  {sheet_note}")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    inject_css()
    init_state()

    if not st.session_state.logged_in:
        page_login()
        return

    render_header()
    render_sidebar()

    if st.session_state.results:
        tab_upload, tab_results = st.tabs(["📂 Upload / Run", "📊 Results"])
        with tab_upload:
            page_upload()
        with tab_results:
            page_results()
    else:
        page_upload()


if __name__ == "__main__":
    main()
